#!/usr/bin/env python3
"""
Ohio SOS CFOFS upload exporter (the "compliance" module).

Reads a working spreadsheet (.xls / .xlsx) with three tabs -- Contributions,
Expense, Loan -- and writes three import-ready, header-less CSV files that load
into the Ohio Secretary of State Campaign Finance Online Filing System (CFOFS)
via  Other Tasks -> Upload Transaction Files.

  Contributions -> {ENTITY}_CONT_{REPORT}.csv   (Schedule 31-A)
  Expense       -> {ENTITY}_EXPS_{REPORT}.csv   (Schedule 31-B)
  Loan          -> {ENTITY}_LOAN_{REPORT}.csv   (Schedule 31-N / 31-C)

CFOFS validates 31-A, 31-B and 31-N/31-C in separate upload steps, so the three
schedules are never combined into one file. The exported CSV has NO header row;
the header exists only in the source workbook for human editing and is stripped
on export.

Usage
-----
    python scripts/compliance/cfofs_export.py path/to/workbook.xlsx
    python scripts/compliance/cfofs_export.py workbook.xls --config myconfig.yaml
    python scripts/compliance/cfofs_export.py workbook.xlsx \
        --entity 16372 --report PostPrimary2026 --output-dir ./out

Config (config.yaml next to this script, overridable per flag):
    entity:                    "16372"            # filenames only, never a column
    report:                    "PostPrimary2026"  # filenames only
    output_dir:                "./out"
    form_of_contribution_code: "4"                # VERIFY via CFOFS Data Download

The tool prints a validation report (per-file row counts, AMOUNT totals, and any
blocking errors with row numbers) BEFORE writing. It refuses to write a file
that has blocking errors -- the clean files are still written and the held file
is named in the report. Exits non-zero if any file is held or a tab is missing.

Dependencies
------------
    pip install openpyxl   # for .xlsx
    pip install xlrd       # for .xls   (xlrd >= 2.0 reads .xls)
    PyYAML is optional; a flat "key: value" config parser is used as a fallback.
"""

import argparse
import csv
import os
import re
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

# --------------------------------------------------------------------------- #
# Column schemas -- exact header strings and order from the CFOFS template.
# The headers are NOT written to the CSV; they document the fixed column order
# and let us generate a blank template. The parallel *_IDS lists are the short
# field keys we read each source row into, positionally.
# --------------------------------------------------------------------------- #

CONT_HEADERS = [
    "CONTRIBUTOR FIRST NAME", "CONTRIB MIDDLE NAME", "CONTRIBUTOR LAST NAME",
    "SUFFIX", "NON INDIVIDUAL CONTRIBUTOR", "PAC REGISTRATION NUMBER",
    "CONTRIBUTOR ADDRESS", "CITY", "STATE", "ZIP",
    "CONTRIBUTOR EMPLOYER OCCUPATION OR LABOR ORGANIZATION",
    "FORM OF CONTRIBUTION", "DATE OF CONTRIBUTION", "AMOUNT",
    "OTHER INCOME TYPE", "EVENT DATE", "INKIND DESCRIPTION",
    "RECEIVED AT FUNDRAISING EVENT (Y/N)", "NAME OF CREDITOR",
    "AMOUNT OF DEBT REMAINING", "ITEM NUMBER", "SCHEDULE CODE (EX. 31A)",
]
CONT_IDS = [
    "first", "middle", "last", "suffix", "non_individual", "pac_reg",
    "address", "city", "state", "zip", "employer", "form", "date", "amount",
    "other_income_type", "event_date", "inkind_desc", "received_at_event",
    "name_of_creditor", "amount_debt_remaining", "item_number", "schedule_code",
]

EXPS_HEADERS = [
    "PAYEE FIRST NAME", "PAYEE MIDDLE NAME", "PAYEE LAST NAME", "SUFFIX",
    "NON INDIVIDUAL PAYEE", "PAYEE ADDRESS", "CITY", "STATE", "ZIP",
    "DATE OF EXPENDITURE", "AMOUNT", "PURPOSE", "EVENT DATE",
    "CANDIDATE OR BALLOT ISSUE (FORM 31U ONLY)",
    "SUPPORT/ OPPOSE (1 if Support/2 if Oppose)", "OFFICE (FORM 31U ONLY)",
    "EXPENDITURE FROM POLITICAL PARTY FUND (FORM 31M ONLY)",
    "ITEM NUMBER", "SCHEDULE CODE (EX. 31B)",
]
EXPS_IDS = [
    "first", "middle", "last", "suffix", "non_individual", "address", "city",
    "state", "zip", "date", "amount", "purpose", "event_date",
    "candidate_or_issue", "support_oppose", "office", "party_fund",
    "item_number", "schedule_code",
]

# NOTE: column 5 header is misspelled "NON INDIDIVUAL" in the official template.
# Keep it exactly -- CFOFS matches on position, but the spelling is preserved so
# anyone regenerating a template from this list gets the official wording.
LOAN_HEADERS = [
    "FIRST NAME", "MIDDLE NAME", "LAST NAME", "SUFFIX", "NON INDIDIVUAL",
    "PAC REGISTRATION NUMBER", "ADDRESS", "CITY", "STATE", "ZIP",
    "EMPLOYER OCCUPATION LABOR ORGANIZATION",
    "DATE LOAN WAS ORIGINAL INCURRED", "PRIOR AMOUNT", "OUTSTANDING BALANCE",
    "PURPOSE", "FORGIVEN (If True Enter a 1)", "AMOUNT INCURRED",
    "PAYMENT DATE", "PAYMENT AMOUNT", "ITEM NUMBER",
    "SCHEDULE CODE (EX 31C, 31N)",
]
LOAN_IDS = [
    "first", "middle", "last", "suffix", "non_individual", "pac_reg",
    "address", "city", "state", "zip", "employer", "date_incurred",
    "prior_amount", "outstanding_balance", "purpose", "forgiven",
    "amount_incurred", "payment_date", "payment_amount", "item_number",
    "schedule_code",
]

# FORM OF CONTRIBUTION: display labels that must be mapped to the SOS code.
# 1 = Check, 2 = Cash, 3 = Credit Card, 4 = Electronic / Electronic Transfer.
# VERIFY the legend via CFOFS Data Download before filing; override the default
# blank-fill via config "form_of_contribution_code".
FORM_LABEL_TO_CODE = {
    "check": "1", "cheque": "1",
    "cash": "2",
    "credit card": "3", "creditcard": "3", "credit": "3", "card": "3",
    "electronic": "4", "electronic transfer": "4", "electronic funds transfer": "4",
    "ach": "4", "eft": "4", "actblue": "4", "wire": "4", "online": "4",
}

# Address values that look like a real entry but are placeholders -- a missing
# creditor street address is a blocking error on the Loan tab.
PLACEHOLDER_ADDRESSES = {
    "tbd", "n/a", "na", "none", "unknown", "-", "--", ".", "pending",
    "address", "on file", "see notes", "xxx", "x",
}

DEFAULT_CONFIG = {
    "entity": "",
    "report": "",
    "output_dir": "./out",
    "form_of_contribution_code": "4",
    # When true, the filer's own entity number is stamped into the PAC
    # REGISTRATION NUMBER column on every Contributions/Loan row (a row that
    # already carries a different PAC number -- a contributor that is itself a
    # PAC -- is left as-is). Note: the CFOFS template guidance says the filer's
    # own number does not belong in this column (the system tags the filer from
    # the logged-in session); this is enabled deliberately per filer request.
    "stamp_filer_pac_number": "true",
}


# --------------------------------------------------------------------------- #
# Config loading
# --------------------------------------------------------------------------- #

def load_config(path):
    """Load config from YAML (PyYAML) or a flat 'key: value' fallback parser."""
    cfg = dict(DEFAULT_CONFIG)
    if not path or not os.path.exists(path):
        return cfg
    text = open(path, encoding="utf-8").read()
    parsed = None
    try:
        import yaml  # type: ignore
        parsed = yaml.safe_load(text) or {}
    except ImportError:
        parsed = _flat_yaml(text)
    if isinstance(parsed, dict):
        for key in DEFAULT_CONFIG:
            if parsed.get(key) is not None:
                cfg[key] = str(parsed[key]).strip()
    return cfg


def _flat_yaml(text):
    """Minimal 'key: value' parser for the flat config (no PyYAML needed)."""
    out = {}
    for line in text.splitlines():
        line = line.split("#", 1)[0].rstrip()
        if not line or ":" not in line:
            continue
        key, _, val = line.partition(":")
        val = val.strip().strip('"').strip("'")
        out[key.strip()] = val
    return out


# --------------------------------------------------------------------------- #
# Field normalizers
# --------------------------------------------------------------------------- #

def collapse(value):
    """Trim and collapse internal runs of whitespace to a single space."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def proper_case(value):
    """Proper Case names/cities -- but only rewrite ALL-CAPS / caseless input.

    If the human already typed mixed case (McDonald, O'Brien), trust it.
    """
    s = collapse(value)
    if not s or any(c.islower() for c in s):
        return s
    return re.sub(
        r"[A-Za-z]+",
        lambda m: m.group()[:1].upper() + m.group()[1:].lower(),
        s,
    )


def normalize_zip(value):
    """ZIP -> first 5 digits. Returns (zip5, error_or_None)."""
    s = collapse(value)
    if not s:
        return "", None
    digits = re.sub(r"\D", "", s)
    if len(digits) < 5:
        return s, f"ZIP '{s}' is not 5 digits"
    return digits[:5], None


def parse_amount(value):
    """Parse a money value. Returns (Decimal_or_None, error_or_None).

    Blank -> (None, None). Strips $ , and spaces; parentheses or a leading
    minus mark a negative (which the schedules forbid, flagged by the caller).
    """
    if value is None:
        return None, None
    if isinstance(value, bool):
        return None, f"amount '{value}' is not a number"
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01")), None
        except InvalidOperation:
            return None, f"unparseable amount '{value}'"
    s = collapse(value)
    if not s:
        return None, None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative, s = True, s[1:-1]
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("-"):
        negative, s = True, s[1:]
    try:
        amount = Decimal(s).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None, f"unparseable amount '{value}'"
    if negative:
        amount = -amount
    return amount, None


def fmt_amount(amount):
    """Plain decimal, two places, no $/commas/parens."""
    return f"{amount:.2f}"


def parse_date(value):
    """Parse a date to MM/DD/YYYY. Returns (text_or_None, error_or_None).

    Handles datetime/date objects (xlsx & converted xls cells), Excel date
    serials (a number like 46128 -> 04/16/2026, 1900 epoch), and text dates.
    """
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y"), None
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y"), None
    if isinstance(value, bool):
        return None, f"invalid date '{value}'"
    if isinstance(value, (int, float)):
        return _serial_to_date(value)
    s = collapse(value)
    if not s:
        return None, None
    # A bare integer string is almost certainly an Excel serial.
    if re.fullmatch(r"\d+(\.\d+)?", s):
        return _serial_to_date(float(s))
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y",
                "%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%B %d, %Y",
                "%b %d, %Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%m/%d/%Y"), None
        except ValueError:
            continue
    return None, f"unparseable date '{s}'"


def _serial_to_date(serial):
    """Excel 1900-system serial -> MM/DD/YYYY. Base 1899-12-30 absorbs the
    Excel 1900-leap-year bug for serials >= 61 (the common range)."""
    if serial <= 0:
        return None, f"invalid date serial '{serial}'"
    try:
        dt = datetime(1899, 12, 30) + timedelta(days=float(serial))
    except (OverflowError, ValueError):
        return None, f"invalid date serial '{serial}'"
    return dt.strftime("%m/%d/%Y"), None


def normalize_form(value, default_code):
    """FORM OF CONTRIBUTION -> SOS code. Returns (code_or_text, error_or_None).

    Blank -> default code. Already-numeric -> passthrough. Known label -> code.
    Anything else stays as the original text and is flagged (CFOFS rejects text
    labels like ELECTRONIC TRANSFER on upload).
    """
    s = collapse(value)
    if not s:
        return default_code, None
    if s.isdigit():
        return s, None
    key = s.lower()
    if key in FORM_LABEL_TO_CODE:
        return FORM_LABEL_TO_CODE[key], None
    return s, f"FORM OF CONTRIBUTION '{s}' is a text label, not an SOS code"


def normalize_yesno(value, default="N"):
    s = collapse(value).lower()
    if not s:
        return default
    return "Y" if s in ("y", "yes", "true", "1") else "N"


def normalize_forgiven(value):
    s = collapse(value).lower()
    return "1" if s in ("1", "y", "yes", "true", "forgiven") else ""


def is_truthy(value):
    return collapse(value).lower() in ("1", "y", "yes", "true", "on")


def filer_pac_number(fields, cfg):
    """PAC REGISTRATION NUMBER for a row. A real value from the source (a
    contributor/creditor that is itself a PAC) wins; otherwise the filer's own
    entity number is stamped in when stamp_filer_pac_number is enabled."""
    existing = collapse(fields.get("pac_reg"))
    if existing:
        return existing
    if is_truthy(cfg.get("stamp_filer_pac_number")) and cfg.get("entity"):
        return cfg["entity"]
    return ""


# --------------------------------------------------------------------------- #
# Workbook reading (backend-agnostic): {sheet_name: [[cell, ...], ...]}
# --------------------------------------------------------------------------- #

def read_workbook(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    if ext == ".xls":
        return _read_xls(path)
    raise SystemExit(f"Unsupported file type '{ext}'. Use .xls or .xlsx.")


def _read_xlsx(path):
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise SystemExit("Missing dependency for .xlsx. Install:  pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = {}
    for ws in wb.worksheets:
        sheets[ws.title] = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return sheets


def _read_xls(path):
    try:
        import xlrd  # type: ignore
    except ImportError:
        raise SystemExit("Missing dependency for .xls. Install:  pip install xlrd")
    book = xlrd.open_workbook(path)
    sheets = {}
    for sheet in book.sheets():
        rows = []
        for ri in range(sheet.nrows):
            row = []
            for ci in range(sheet.ncols):
                cell = sheet.cell(ri, ci)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                row.append(value)
            rows.append(row)
        sheets[sheet.name] = rows
    return sheets


def match_tabs(sheets):
    """Map the three required tabs case/space-insensitively. Returns a dict of
    schedule -> (sheet_name, rows) or None when the tab is absent."""
    found = {"CONT": None, "EXPS": None, "LOAN": None}
    for name, rows in sheets.items():
        key = re.sub(r"\s+", "", name.strip().lower())
        if "contribution" in key and found["CONT"] is None:
            found["CONT"] = (name, rows)
        elif ("expens" in key or "expenditure" in key) and found["EXPS"] is None:
            found["EXPS"] = (name, rows)
        elif ("loan" in key or "debt" in key) and found["LOAN"] is None:
            found["LOAN"] = (name, rows)
    return found


def is_blank_row(row):
    return all(c is None or str(c).strip() == "" for c in row)


def data_rows(rows):
    """Strip the header row (row 1) and drop fully-empty rows.

    Yields (spreadsheet_row_number, row_cells). Row numbers are 1-based with the
    header at row 1, so the first data row reports as row 2.
    """
    for idx, row in enumerate(rows[1:], start=2):
        if is_blank_row(row):
            continue
        yield idx, row


def row_fields(row, ids):
    return {fid: (row[i] if i < len(row) else None) for i, fid in enumerate(ids)}


# --------------------------------------------------------------------------- #
# Shared validation helpers
# --------------------------------------------------------------------------- #

def check_person_or_org(fields, errors, label):
    """Enforce the name-XOR-organization rule. Returns True if a person."""
    name_filled = bool(collapse(fields.get("first")) or collapse(fields.get("last")))
    any_name_part = any(
        collapse(fields.get(k)) for k in ("first", "middle", "last", "suffix")
    )
    org_filled = bool(collapse(fields.get("non_individual")))
    if org_filled and any_name_part:
        errors.append(f"both individual name and NON INDIVIDUAL {label} are filled")
    elif not name_filled and not org_filled:
        errors.append(f"missing {label}: no name parts and no NON INDIVIDUAL value")
    return name_filled and not org_filled


def check_address(fields, errors, label, flag_placeholder=False):
    addr = collapse(fields.get("address"))
    city = collapse(fields.get("city"))
    state = collapse(fields.get("state"))
    zip_value, zip_err = normalize_zip(fields.get("zip"))
    if not addr:
        errors.append(f"missing {label} street address")
    elif flag_placeholder and addr.lower() in PLACEHOLDER_ADDRESSES:
        errors.append(f"placeholder {label} address '{addr}' is not allowed")
    if not city:
        errors.append("missing city")
    if not state:
        errors.append("missing state")
    if not collapse(fields.get("zip")):
        errors.append("missing ZIP")
    elif zip_err:
        errors.append(zip_err)
    return zip_value


def required_amount(fields, key, errors, label):
    """Parse an amount that must be present and > 0. Returns Decimal or None."""
    amount, err = parse_amount(fields.get(key))
    if err:
        errors.append(err)
        return None
    if amount is None:
        errors.append(f"missing {label}")
        return None
    if amount <= 0:
        errors.append(f"{label} must be greater than 0 (got {fmt_amount(amount)})")
        return None
    return amount


def optional_nonneg_amount(fields, key, errors, label):
    """Parse an optional amount that, if present, must be >= 0. Default 0."""
    amount, err = parse_amount(fields.get(key))
    if err:
        errors.append(err)
        return None
    if amount is None:
        return Decimal("0.00")
    if amount < 0:
        errors.append(f"{label} cannot be negative (got {fmt_amount(amount)})")
        return None
    return amount


# --------------------------------------------------------------------------- #
# Per-schedule processors
# Each returns a list of result dicts:
#   {row, cells, errors, warnings, amount}
# --------------------------------------------------------------------------- #

def process_contributions(rows, cfg):
    results = []
    item = 0
    for rownum, raw in data_rows(rows):
        f = row_fields(raw, CONT_IDS)
        errors, warnings = [], []

        is_person = check_person_or_org(f, errors, "contributor")
        zip_value = check_address(f, errors, "contributor")

        employer = collapse(f.get("employer"))
        if is_person and not employer:
            warnings.append("missing employer (CFOFS flags blank employer for individuals)")

        form_code, form_err = normalize_form(f.get("form"), cfg["form_of_contribution_code"])
        if form_err:
            errors.append(form_err)

        date_text, date_err = parse_date(f.get("date"))
        if date_err:
            errors.append(date_err)
        elif not date_text:
            errors.append("missing DATE OF CONTRIBUTION")

        amount = required_amount(f, "amount", errors, "AMOUNT")

        pac_reg = filer_pac_number(f, cfg)

        event_date_text, ev_err = parse_date(f.get("event_date"))
        if ev_err:
            errors.append(ev_err)

        item += 1
        cells = [
            proper_case(f.get("first")), proper_case(f.get("middle")),
            proper_case(f.get("last")), collapse(f.get("suffix")),
            collapse(f.get("non_individual")), pac_reg,
            collapse(f.get("address")), proper_case(f.get("city")),
            collapse(f.get("state")).upper(), zip_value, employer,
            form_code, date_text or "", fmt_amount(amount) if amount is not None else "",
            collapse(f.get("other_income_type")), event_date_text or "",
            collapse(f.get("inkind_desc")),
            normalize_yesno(f.get("received_at_event"), "N"),
            proper_case(f.get("name_of_creditor")),
            _passthrough_amount(f.get("amount_debt_remaining")),
            str(item), "31A",
        ]
        results.append({"row": rownum, "cells": cells, "errors": errors,
                        "warnings": warnings, "amount": amount or Decimal("0.00")})
    return results


def process_expenses(rows, cfg):
    results = []
    item = 0
    for rownum, raw in data_rows(rows):
        f = row_fields(raw, EXPS_IDS)
        errors, warnings = [], []

        check_person_or_org(f, errors, "payee")
        zip_value = check_address(f, errors, "payee")

        date_text, date_err = parse_date(f.get("date"))
        if date_err:
            errors.append(date_err)
        elif not date_text:
            errors.append("missing DATE OF EXPENDITURE")

        amount = required_amount(f, "amount", errors, "AMOUNT")

        if not collapse(f.get("purpose")):
            warnings.append("missing PURPOSE (CFOFS expects a short, specific purpose)")

        event_date_text, ev_err = parse_date(f.get("event_date"))
        if ev_err:
            errors.append(ev_err)

        item += 1
        cells = [
            proper_case(f.get("first")), proper_case(f.get("middle")),
            proper_case(f.get("last")), collapse(f.get("suffix")),
            collapse(f.get("non_individual")), collapse(f.get("address")),
            proper_case(f.get("city")), collapse(f.get("state")).upper(),
            zip_value, date_text or "",
            fmt_amount(amount) if amount is not None else "",
            collapse(f.get("purpose")), event_date_text or "",
            collapse(f.get("candidate_or_issue")),
            collapse(f.get("support_oppose")), collapse(f.get("office")),
            collapse(f.get("party_fund")), str(item), "31B",
        ]
        results.append({"row": rownum, "cells": cells, "errors": errors,
                        "warnings": warnings, "amount": amount or Decimal("0.00")})
    return results


def process_loans(rows, cfg):
    results = []
    item = 0
    for rownum, raw in data_rows(rows):
        f = row_fields(raw, LOAN_IDS)
        errors, warnings = [], []

        check_person_or_org(f, errors, "creditor")
        zip_value = check_address(f, errors, "creditor", flag_placeholder=True)

        date_text, date_err = parse_date(f.get("date_incurred"))
        if date_err:
            errors.append(date_err)
        elif not date_text:
            errors.append("missing DATE LOAN WAS ORIGINAL INCURRED")

        prior = optional_nonneg_amount(f, "prior_amount", errors, "PRIOR AMOUNT")
        incurred = optional_nonneg_amount(f, "amount_incurred", errors, "AMOUNT INCURRED")
        payment = optional_nonneg_amount(f, "payment_amount", errors, "PAYMENT AMOUNT")

        outstanding_raw, out_err = parse_amount(f.get("outstanding_balance"))
        if out_err:
            errors.append(out_err)
        outstanding = None
        if None not in (prior, incurred, payment):
            expected = (prior + incurred - payment).quantize(Decimal("0.01"))
            if outstanding_raw is None:
                outstanding = expected
            elif outstanding_raw != expected:
                errors.append(
                    f"OUTSTANDING BALANCE {fmt_amount(outstanding_raw)} != prior "
                    f"{fmt_amount(prior)} + incurred {fmt_amount(incurred)} - payment "
                    f"{fmt_amount(payment)} = {fmt_amount(expected)}")
            else:
                outstanding = outstanding_raw
            if outstanding is not None and outstanding < 0:
                errors.append(f"OUTSTANDING BALANCE is negative ({fmt_amount(outstanding)})")

        payment_date_text, pay_err = parse_date(f.get("payment_date"))
        if pay_err:
            errors.append(pay_err)
        if payment and payment > 0 and not payment_date_text:
            warnings.append("PAYMENT AMOUNT present but PAYMENT DATE is blank")

        schedule_code = collapse(f.get("schedule_code")).upper() or "31N"
        if schedule_code not in ("31N", "31C"):
            errors.append(f"SCHEDULE CODE '{schedule_code}' must be 31N (debt) or 31C (loan)")

        pac_reg = filer_pac_number(f, cfg)

        item += 1
        cells = [
            proper_case(f.get("first")), proper_case(f.get("middle")),
            proper_case(f.get("last")), collapse(f.get("suffix")),
            collapse(f.get("non_individual")), pac_reg,
            collapse(f.get("address")), proper_case(f.get("city")),
            collapse(f.get("state")).upper(), zip_value,
            collapse(f.get("employer")), date_text or "",
            fmt_amount(prior) if prior is not None else "",
            fmt_amount(outstanding) if outstanding is not None else "",
            collapse(f.get("purpose")), normalize_forgiven(f.get("forgiven")),
            fmt_amount(incurred) if incurred is not None else "",
            payment_date_text or "",
            fmt_amount(payment) if payment is not None and payment > 0 else "",
            str(item), schedule_code,
        ]
        results.append({"row": rownum, "cells": cells, "errors": errors,
                        "warnings": warnings,
                        "amount": outstanding or Decimal("0.00")})
    return results


def _passthrough_amount(value):
    """Format an optional money cell, leaving blanks blank (no validation)."""
    amount, err = parse_amount(value)
    if err or amount is None:
        return collapse(value)
    return fmt_amount(amount)


# --------------------------------------------------------------------------- #
# Reporting & writing
# --------------------------------------------------------------------------- #

SCHEDULES = [
    ("CONT", "31A", "Contributions", process_contributions),
    ("EXPS", "31B", "Expense", process_expenses),
    ("LOAN", "31N/31C", "Loan", process_loans),
]


def write_csv(path, results):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
        for result in results:
            writer.writerow(result["cells"])


def run(args):
    cfg = load_config(args.config)
    for key in ("entity", "report", "output_dir", "form_of_contribution_code"):
        override = getattr(args, key, None)
        if override:
            cfg[key] = override

    missing_cfg = [k for k in ("entity", "report") if not cfg[k]]
    if missing_cfg:
        sys.stderr.write(
            "Missing required config for filenames: " + ", ".join(missing_cfg) +
            "\nSet them in the config file or pass --entity / --report.\n")
        return 2

    if not os.path.exists(args.workbook):
        sys.stderr.write(f"Workbook not found: {args.workbook}\n")
        return 2

    sheets = read_workbook(args.workbook)
    tabs = match_tabs(sheets)

    print("=" * 72)
    print(f"CFOFS export  |  entity {cfg['entity']}  |  report {cfg['report']}")
    print(f"Source: {args.workbook}")
    print(f"Form-of-contribution default code: {cfg['form_of_contribution_code']} "
          f"(VERIFY via CFOFS Data Download)")
    print("=" * 72)

    files = []          # (schedule, path, clean_results)
    held = []           # schedules held back due to blocking errors
    totals = {}
    exit_code = 0

    for schedule, code, tabname, processor in SCHEDULES:
        print(f"\n[{schedule}]  {tabname}  (Schedule {code})")
        if tabs[schedule] is None:
            print(f"  ! Tab not found in workbook -- expected a '{tabname}' tab.")
            exit_code = 1
            continue
        sheet_name, rows = tabs[schedule]
        results = processor(rows, cfg)

        blocking = [r for r in results if r["errors"]]
        warned = [r for r in results if r["warnings"]]
        total = sum((r["amount"] for r in results), Decimal("0.00"))
        totals[schedule] = total
        amount_label = "outstanding" if schedule == "LOAN" else "AMOUNT"

        print(f"  tab: '{sheet_name}'   rows: {len(results)}   "
              f"total {amount_label}: {fmt_amount(total)}")

        for r in blocking:
            for msg in r["errors"]:
                print(f"  ERROR  row {r['row']}: {msg}")
        for r in warned:
            for msg in r["warnings"]:
                print(f"  warn   row {r['row']}: {msg}")

        filename = f"{cfg['entity']}_{schedule}_{cfg['report']}.csv"
        path = os.path.join(cfg["output_dir"], filename)

        if blocking:
            held.append((schedule, filename, len(blocking)))
            exit_code = 1
            print(f"  HELD: {len(blocking)} row(s) with blocking errors -- "
                  f"{filename} NOT written.")
        elif not results:
            print(f"  0 rows -- nothing to upload. {filename} NOT written "
                  f"(verify there are genuinely no {tabname.lower()} this period).")
        else:
            files.append((schedule, path, results))

    print("\n" + "-" * 72)
    print("Cross-foot (tie these to the cover page before submitting):")
    print(f"  Contributions (31-A) total .......... {fmt_amount(totals.get('CONT', Decimal('0.00')))}")
    print(f"  Expenditures  (31-B) total .......... {fmt_amount(totals.get('EXPS', Decimal('0.00')))}")
    print(f"  Loans/debts   (31-N/C) outstanding .. {fmt_amount(totals.get('LOAN', Decimal('0.00')))}")
    print("-" * 72)

    for schedule, path, results in files:
        write_csv(path, results)
        print(f"WROTE  {path}  ({len(results)} rows, no header)")

    if held:
        print("\nHELD (fix the blocking errors above, then re-run):")
        for schedule, filename, count in held:
            print(f"  {filename}  ({count} row(s) blocked)")

    if not files and not held:
        print("\nNothing written.")

    return exit_code


def main():
    ap = argparse.ArgumentParser(
        description="Export Ohio SOS CFOFS upload CSVs from a 3-tab workbook.")
    ap.add_argument("workbook", help="Path to the source .xls/.xlsx workbook")
    default_config = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.yaml")
    ap.add_argument("--config", default=default_config,
                    help="Path to the YAML config (default: config.yaml beside this script)")
    ap.add_argument("--output-dir", dest="output_dir", help="Override output_dir")
    ap.add_argument("--entity", help="Override entity (filenames only)")
    ap.add_argument("--report", help="Override report (filenames only)")
    ap.add_argument("--form-code", dest="form_of_contribution_code",
                    help="Override the default FORM OF CONTRIBUTION code")
    args = ap.parse_args()
    sys.exit(run(args))


if __name__ == "__main__":
    main()
